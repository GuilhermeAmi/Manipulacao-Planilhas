import openpyxl
import pandas as pd
import re

def remove_space_substrings(line):
    # Substitua todas as substrings indesejadas
    substrings_to_remove = ["  "]
    for substring in substrings_to_remove:
        line = line.replace(substring, ' ').strip()
    return line

def same_combinations_cm(line):
    # Expressão regular para capturar o ângulo retangular
    cm_combinations = r'(\d+CM|\d+,\d+M|\d+,\d+CM|\d+/\d+/\d+/\d+M|\d+X\d+CM|\d+X\d+X\d+CM|\bMULTHI ROUND|\bRGB)'
    
    
    pattern_cm = ''  # Definindo a variável pattern_text com valor vazio

    cm_combinations = re.search(cm_combinations, line)
    
    if cm_combinations:
        pattern_cm = cm_combinations.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(pattern_cm, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\d+,\d+W|\d+W|\d+W/M'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + pattern_cm + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line


def modification_order(line):
    # Expressão regular para capturar o ângulo retangular
    padrao_rgb = r'\bC/ CONTROLE'
     
    rgb_pattern = ''  # Definindo a variável pattern_text com valor vazio

    padrao_rgb = re.search(padrao_rgb, line)
    
    if padrao_rgb:
        rgb_pattern = padrao_rgb.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(rgb_pattern, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\d+,\d+W|\d+W'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + rgb_pattern + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line

def modification_rgb(line):
    # Expressão regular para capturar o ângulo retangular
    padrao_rgb = r'\bRGB'
     
    rgb_pattern = ''  # Definindo a variável pattern_text com valor vazio

    padrao_rgb = re.search(padrao_rgb, line)
    
    if padrao_rgb:
        rgb_pattern = padrao_rgb.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(rgb_pattern, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\d+,\d+W|\d+W'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + rgb_pattern + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line


def modification_soquet(line):
    # Expressão regular para capturar o ângulo retangular
    modification_led = r'E(\d+)'
    
    
    led_pattern = ''  # Definindo a variável pattern_text com valor vazio

    modification_led = re.search(modification_led, line)
    
    if modification_led:
        led_pattern = modification_led.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(led_pattern, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\d+,\d+W|\d+W'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + led_pattern + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line


def modification_smart(line):
    # Expressão regular para capturar o ângulo retangular
    modification_led = r'\bSMART'
    
    
    led_pattern = ''  # Definindo a variável pattern_text com valor vazio

    modification_led = re.search(modification_led, line)
    
    if modification_led:
        led_pattern = modification_led.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(led_pattern, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\d+,\d+W|\d+W'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + led_pattern + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line



def modification_diretions(line):
    # Expressão regular para capturar o ângulo retangular
    modification_led = r'\bRETANG$|\bRETANGULAR|\bRED$|\bREDONDO|\bDIM$'
    
    
    led_pattern = ''  # Definindo a variável pattern_text com valor vazio

    modification_led = re.search(modification_led, line)
    
    if modification_led:
        led_pattern = modification_led.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(led_pattern, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\d+,\d+W|\d+W'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + led_pattern + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line



def modification_led(line):
    # Expressão regular para capturar o ângulo retangular
    modification_led = r'(\b\d+LEDS/M\b|\d+LEDS)'
    
    
    led_pattern = ''  # Definindo a variável pattern_text com valor vazio

    modification_led = re.search(modification_led, line)
    
    if modification_led:
        led_pattern = modification_led.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(led_pattern, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\dccc|\d+W|\d+W/M|\d+,\d+W|\d+.\d+W'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + led_pattern + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line


def preprocess_fachos(line):
    # Expressão regular para capturar o ângulo retangular
    fachos_combinations = r'\d+\sFACHOS|\d+\sFACHO\s\w+|\d+\sFACHO'
    
    
    pattern_fachos = ''  # Definindo a variável pattern_text com valor vazio

    fachos_combinations = re.search(fachos_combinations, line)
    
    if fachos_combinations:
        pattern_fachos = fachos_combinations.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(pattern_fachos, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\d+,\d+W|\d+W'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + pattern_fachos + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line



def remove_unwanted_substrings(line):
    # Substitua todas as substrings indesejadas
    substrings_to_remove = ["-", "(LIQUIDACAO)", "(LIQUI)", "(FF)"]
    for substring in substrings_to_remove:
        line = line.replace(substring, '').strip()
    return line


def modification_degree(line):
    # Expressão regular para capturar o ângulo retangular
    modification_led = r'(\d+°\s\w+|\d+°\w+|\d+°)'
    
    
    led_pattern = ''  # Definindo a variável pattern_text com valor vazio

    modification_led = re.search(modification_led, line)
    
    if modification_led:
        led_pattern = modification_led.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(led_pattern, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\d+,\d+W|\d+W'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + led_pattern + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line


def modification_control(line):
    # Expressão regular para capturar o ângulo retangular
    modification_led = r'\bC/CONTROLE$|\bC/ CONTROLE'
    
    
    led_pattern = ''  # Definindo a variável pattern_text com valor vazio

    modification_led = re.search(modification_led, line)
    
    if modification_led:
        led_pattern = modification_led.group()
        
    # Remove o padrão capturado da linha original
    line_without_pattern = line.replace(led_pattern, '').strip()

    # Encontra o índice do primeiro valor de watts
    watts_pattern = r'\d+,\d+W|\d+W'
    watts_match = re.search(watts_pattern, line)

    if watts_match:
        index_watts = watts_match.start()

    # Insere o padrão capturado no final do Modelo e rearranja a linha
        new_line = line_without_pattern[:index_watts] + led_pattern + ' ' + line_without_pattern[index_watts:]
        return new_line

    return line



def extract_data_from_line(line):
    # Expressão regular para capturar os diferentes campos do modelo
    pattern1 = r'^(\d+) (.+?) (\d+,\d+W|\d+W|\d+,\d+W/M) (\d+K|\d+/\d+/\d+K|\d+/\d+K|\d+K/\d+K/\d+K|\d+K/\d+K) ((?:[\w\s]+/)*[\w\s]+)(?: (\w+(?:\s\d+XE-\d+)?))?$'
    match = re.match(pattern1, line)

    pattern2 = r'^(\d+) (.+?) (\w+/\w+|\w+) (\d+XE\d+|\d+XG\d+|\d+XGU\d+|\d+XT\d+|\d+XMR\d+|\d+XDICROICA|\d+XMINI\s\bDICROICA)'
    new_match = re.match(pattern2, line)

    pattern3 = r'^(\d+) (.+?) (\w+/\w+|\w+) (\d+CM) (\d+XE\d+|\d+XG\d+|\d+XMR\d+)'
    new_match1 = re.match(pattern3, line)

    pattern4 = r'^(\d+) (.+?) (\b\d+LEDS/M\b) (\d+,\d+W/M|\d+W/M) (\d+K) (\d+V)  (\d+LM/M) (IP\d+)'
    new_match4 = re.match(pattern4, line)

    pattern5 = r'^(\d+) (\bFIITA\s\bLED) (\d+LEDS|\b\d+LEDS/M\b) (\d+W/M|\d+,\d+W/M|\d+.\d+W) (\d+K) (\d+V)  (\d+LM/M) (\wROLO\s\d+M)$'
    new_match5 = re.match(pattern5, line)



    if match:
        # Captura os grupos da expressão regular somente se eles existirem na string
        sku = match.group(1)
        modelo = match.group(2)
        watts = match.group(3) if match.group(3) else None
        temperatura = match.group(4) if match.group(4) else None
        cor_fluxo = match.group(5) if match.group(5) else None
        exemplo = line 
        
        if cor_fluxo:
            cor_fluxo = cor_fluxo.split()
            cor = ' '.join([item for item in cor_fluxo if not item.endswith('LM') and not item.startswith('IP')])
            fluxo = next((item for item in cor_fluxo if item.endswith('LM')), None)
            # capturar IP caso haja na linha
            ip_pattern = r'IP(\d+)'
            ip_match = re.search(ip_pattern, ' '.join(cor_fluxo))
            ip = 'IP' + ip_match.group(1) if ip_match else None
            # capturar a volt   agem caso haja na linha
            voltagem_pattern = r'(BIV)|(\w+VOLT)|(\bAUTOVOLT)|(\d+V$)'
            voltagem_match = re.search(voltagem_pattern, ' '.join(cor_fluxo))
            voltagem = voltagem_match.group(1) if voltagem_match and voltagem_match.group(1) and not voltagem_match.group(1).startswith('VOLT') else None
        else:
            cor = None
            fluxo = None
            voltagem = None
            ip = None

        return sku, modelo, watts, temperatura, cor, voltagem, fluxo, ip, 'Descrição Site', None, "Marca", exemplo
    
    elif new_match1:
        # Caso a linha não corresponda ao padrão, tenta organizar de acordo com um novo padrão
        # Expressão regular para capturar SKU, modelo e descrição
        new_match1 = re.match(pattern3, line)
        
        if new_match:
            sku_line = new_match1.group(1)
            descricao = new_match1.group(2)
            cor_line = new_match1.group(3)
            centimetros = new_match1.group(4)
            soquete = new_match1.group(5)
            exemplo = line

            
            if soquete:
                lamp_pattern = r'\d+X$'
                lamp_match = re.search(lamp_pattern, ' '.join(soquete))
                lamp = lamp_match.group(1) if lamp_match else None
                ####################################################
                soq_pattern = r'\bE(\d+)|\bE-(\d+)|\bGU(\d+)|\bMR(\d+)|\bMINI\s\bDICROICA'
                soq_match = re.search(soq_pattern, ' '.join(soquete))
                soq = soq_match.group(1) if soq_match else None
            
            # Organizar a linha conforme o novo padrão
            soquete = f'{lamp} {soq}'
            descricao = f'{descricao} {centimetros}'
            
            return sku_line, descricao, None, None, cor_line, None, None, None, None, soquete, None, exemplo
        
    elif new_match:
        # Caso a linha não corresponda ao padrão, tenta organizar de acordo com um novo padrão
        # Expressão regular para capturar SKU, modelo e descrição
        new_match = re.match(pattern2, line)
        
        if new_match:
            sku_line = new_match.group(1)
            descricao = new_match.group(2)
            cor_line = new_match.group(3)
            soquete = new_match.group(4)
            exemplo = line

            
            return sku_line, descricao, None, None, cor_line, None, None, None, 'Descrição Site', soquete, "Marca", exemplo

    elif new_match4:
        pattern4 = r'^(\d+) (.+?) (\d+,\d+W/M|\d+W/M) (\d+K) (\d+V) (\d+LEDS/M) (\d+LM/M) (IP\d+)'
        new_match4 = re.match(pattern4, line)

        if new_match4:
            sku_line4 = new_match4.group(1)
            descricao4 = new_match4.group(2)
            watts4 = new_match4.group(3)
            temperatura4 = new_match4.group(4)
            voltagem4 = new_match4.group(5)
            leds = new_match4.group(6)
            fluxo4 = new_match4.group(7)
            # ip4 = new_match4.group(8) if new_match4.group(8) else None
            exemplo = line 
            

            descricao4 = f'{descricao4} {leds}'


            return sku_line4, descricao4, watts4, temperatura4, None, voltagem4, fluxo4, None, None, None, exemplo     

    elif new_match5:
        #pattern5 = r'^(\d+) (\bFITA\s\bLED) (\d+LEDS/M|\d+LEDS) ((\d+)W/M|\d+,\d+W/M|\d+.\d+W) (\d+K/\d+k) (\d+V) (\d+LM/M) (ROLO\s\d+M)$'
        new_match5 = re.match(pattern5, line)

        if new_match5:
            sku5 = new_match5.group(1)
            descricao5 = new_match5.group(2)
            leds5 = new_match5.group(3)
            watts5 = new_match5.group(4)
            temperatura5 = new_match5.group(5)
            voltagem5 = new_match5.group(6)
            fluxo5 = new_match5.group(7)
            tamanho5 = new_match5.group(8)
    
    
            descricao = f'{descricao5} {leds5} {tamanho5}'
    
            return sku5, descricao5, watts5, temperatura5, None, voltagem5, fluxo5, None, 'Descrição Site', None, "Marca", line



with open('produtos_janeiro_2023.txt', 'r', encoding='utf-8') as file: 
        linhas = file.read().splitlines()

data_list = []
# Adicione este bloco de código para depuração
for linha in linhas:
    remover_espacos = remove_space_substrings(linha)
    linha = remove_unwanted_substrings(remover_espacos)
    modicacoes_encomenda = modification_order(linha)
    modificacoes_rgb = modification_rgb(modicacoes_encomenda)
    modificacoes_soquete = modification_soquet(modificacoes_rgb)
    modificacoes_smart = modification_smart(modificacoes_soquete)
    modificacoes_direcoes = modification_diretions(modificacoes_smart)
    modificacoes_led = modification_led(modificacoes_direcoes)
    fachos_alinhados = preprocess_fachos(modificacoes_led)
    linha_processada = remove_unwanted_substrings(fachos_alinhados)
    combinações = same_combinations_cm(linha_processada)
    modificacoes_grau = modification_degree(combinações)
    modificacoes_controle = modification_control(modificacoes_grau)
    raw_data = extract_data_from_line(modificacoes_controle)
    if raw_data:
        data_list.append(raw_data)
    else:
        data_list.append((None, linha, None, None, None, None, None, None, None, None, None, None))
df = pd.DataFrame(data_list, columns=['SKU', 'Modelo', 'Watts', 'Temperatura de Cor', 'Cor', 'Voltagem', 'Fluxo Luminoso', 'IP', 'Descrição Site', 'Soquete', 'Marca', 'Exemplo'])

# Preenche a coluna "Modelo" com a linha original nas linhas onde os dados extraídos estão faltando
df['Modelo'] = df.apply(lambda row: row['Modelo'] if row['Watts'] and row['Temperatura de Cor'] and row['Cor'] and row['Voltagem'] and row['Fluxo Luminoso'] and row['IP'] else row['Modelo'], axis=1)

# Trata a coluna 'Cor' para combinar múltiplas cores separadas por "/"
df['Cor'] = df['Cor'].apply(lambda x: '/'.join(sorted(set(x.split('/')))) if (x and '/' in x) else x)

# Trata a coluna 'Temperatura de Cor' para combinar múltiplas cores separadas por "/"
df['Temperatura de Cor'] = df['Temperatura de Cor'].apply(lambda x: '/'.join(sorted(set(x.split('/')))) if (x and '/' in x) else x)

# Substitui "BIV" por "BIVOLT" na coluna "Voltagem"
df['Voltagem'] = df['Voltagem'].str.replace(r'\bBIV\b', 'BIVOLT', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\bCR\b', 'CROMADO', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\bTRANSP$|\b/TRANSP$', 'TRANSPARENTE', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\bBCA', 'BRANCA', regex=True) 

# Remove "BIV" da coluna "cor"
df['Cor'] = df['Cor'].str.replace(r'\bDR$', 'DOURADO', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\bPTO', 'PRETO', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\bBCO', 'BRANCO', regex=True)

df['Cor'] = df['Cor'].str.replace(r'BIV', '', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\bQUADRADO|\bREDONDO', '', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\d+V', '', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\d+CM', '', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\bCOB', 'COBRE', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\bMAD$', 'MADEIRA', regex=True)

df['Cor'] = df['Cor'].str.replace(r'\bOLT', '', regex=True)

df['Cor'] = df['Cor'].str.replace(r'E(\d+)', '', regex=True)

df['Modelo'] = df["Modelo"].str.replace(r'\bRETANG$|\bRETANG\s', 'RETANGULAR', regex=True)

df['Modelo'] = df["Modelo"].str.replace(r'         ', ' ', regex=True)

df['Modelo'] = df['Modelo'].str.replace(r'\bQUAD$|\bQUAD\s$', 'QUADRADA', regex=True)

df['Modelo'] = df['Modelo'].str.replace(r'\bRED.$|\bRED$', 'REDONDO', regex=True)

df['Modelo'] = df['Modelo'].str.replace(r'\bDIM.', 'DIMERIZÁVEL', regex=True)

df['Soquete'] = df['Soquete'].str.replace(r'XE', 'Lâmpada  E', regex=True)

df['Soquete'] = df['Soquete'].str.replace(r'XG', 'Lâmpada  G', regex=True)

df['Soquete'] = df['Soquete'].str.replace(r'XGU', 'Lâmpada  GU', regex=True)

df['Soquete'] = df['Soquete'].str.replace(r'XMR', 'Lâmpada  MR', regex=True)

df['Fluxo Luminoso'] = df['Fluxo Luminoso'].str.replace(r'\b/M', '', regex=True)






for index, row in df.iterrows():
    formula = f'=procv(a{index + 2};$planilha3.a:c;3;0)'  # O índice é baseado em zero, então adicionamos 2 para corresponder à linha real.
    df.at[index, 'Marca'] = formula.upper()

for index, row in df.iterrows():
    formula_Desc = f'=arrumar(concatenar(B{index + 2};" ";C{index + 2};" ";D{index + 2};" ";E{index + 2};" ";G{index + 2};" ";H{index + 2};" ";F{index + 2};" ";J{index + 2}))'  # O índice é baseado em zero, então adicionamos 2 para corresponder à linha real.
    df.at[index, 'Descrição Site'] = formula_Desc.upper()


# Cria o arquivo Excel
Project = openpyxl.Workbook()
Arandelas_page = Project.create_sheet('Resultados')
Planilha3_page = Project.create_sheet('Planilha3')

# Adiciona os cabeçalhos
Arandelas_page.append(['SKU', 'Modelo', 'Watts', 'Temperatura de Cor', 'Cor', 'Voltagem', 'Fluxo Luminoso', 'IP', 'Descrição Site','Soquete', 'Marca','Exemplo'])
Planilha3_page.append(['idsubproduto', 'descrição', 'marca'])

# Preenche a planilha com os dados do DataFrame
for row in df.itertuples(index=False):
    Arandelas_page.append(list(row))

# Ajusta a largura das colunas na planilha
for col in Arandelas_page.columns:
    max_length = 0
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    Arandelas_page.column_dimensions[col[0].column_letter].width = adjusted_width


# Salva o arquivo Excel
Project.save('produtos_janeiro_2023.xlsx')    